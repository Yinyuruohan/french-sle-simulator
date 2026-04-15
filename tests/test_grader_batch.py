"""
Tests for grader/batch.py — unit tests use mocked DB calls;
integration tests use Flask test client with a real temp DB.
"""
import io
import json
import sqlite3
import uuid
from datetime import datetime

import openpyxl
import pytest


# ── Helpers ──────────────────────────────────────────────────────────────────

def _make_question(qid: int, options=None, explanation=None):
    return {
        "question_id": qid,
        "correct_answer": "A",
        "grammar_topic": "agreement",
        "options": options or {"A": "opt A", "B": "opt B", "C": "opt C", "D": "opt D"},
        "explanation": explanation or {"why_correct": "Reason", "grammar_rule": "Rule"},
    }


def _load_wb(file_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(file_bytes))


def _make_xlsx(rows: list) -> bytes:
    """Build a minimal .xlsx file in memory from a list of rows. First row is the header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_data in rows:
        ws.append(row_data)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Export unit tests ─────────────────────────────────────────────────────────

class TestExportToExcel:
    def setup_method(self):
        from grader.batch import COLUMNS
        self.COLUMNS = COLUMNS

    def _run_export(self, contexts_result, context_data_map, reviews_map):
        from grader.batch import export_to_excel
        return export_to_excel(
            {},
            _get_contexts=lambda filters: contexts_result,
            _get_context_data=lambda cid: context_data_map.get(cid),
            _get_review=lambda cid: reviews_map.get(cid),
        )

    def test_correct_columns_in_header(self):
        """Header row contains all expected column names in order."""
        file_bytes = self._run_export({"items": []}, {}, {})
        wb = _load_wb(file_bytes)
        ws = wb.active
        header = [ws.cell(row=1, column=i).value for i in range(1, len(self.COLUMNS) + 1)]
        assert header == self.COLUMNS

    def test_empty_contexts_produces_header_only(self):
        """No contexts → only header row, no data rows."""
        file_bytes = self._run_export({"items": []}, {}, {})
        wb = _load_wb(file_bytes)
        ws = wb.active
        assert ws.cell(row=2, column=1).value is None

    def test_one_row_per_question(self):
        """Context with 2 questions produces 2 data rows."""
        ctx_data = {
            "context_id": "ctx-1",
            "passage": "Test passage",
            "questions": [_make_question(1), _make_question(2)],
        }
        file_bytes = self._run_export(
            {"items": [{"context_id": "ctx-1"}]},
            {"ctx-1": ctx_data},
            {},
        )
        wb = _load_wb(file_bytes)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "ctx-1"
        assert ws.cell(row=3, column=1).value == "ctx-1"
        assert ws.cell(row=4, column=1).value is None

    def test_context_fields_repeat_across_question_rows(self):
        """passage and context_id repeat on each question row."""
        ctx_data = {
            "context_id": "ctx-2",
            "passage": "My passage",
            "questions": [_make_question(1), _make_question(2)],
        }
        file_bytes = self._run_export(
            {"items": [{"context_id": "ctx-2"}]},
            {"ctx-2": ctx_data},
            {},
        )
        wb = _load_wb(file_bytes)
        ws = wb.active
        passage_col = self.COLUMNS.index("passage") + 1
        assert ws.cell(row=2, column=passage_col).value == "My passage"
        assert ws.cell(row=3, column=passage_col).value == "My passage"

    def test_expert_rating_on_first_row_only(self):
        """expert_rating populated on first question row, blank on subsequent rows."""
        ctx_data = {
            "context_id": "ctx-3",
            "passage": "Passage",
            "questions": [_make_question(1), _make_question(2)],
        }
        review = {"expert_rating": "Good", "expert_critique": "Looks fine."}
        file_bytes = self._run_export(
            {"items": [{"context_id": "ctx-3"}]},
            {"ctx-3": ctx_data},
            {"ctx-3": review},
        )
        wb = _load_wb(file_bytes)
        ws = wb.active
        rating_col = self.COLUMNS.index("expert_rating") + 1
        critique_col = self.COLUMNS.index("expert_critique") + 1
        assert ws.cell(row=2, column=rating_col).value == "Good"
        assert ws.cell(row=2, column=critique_col).value == "Looks fine."
        assert ws.cell(row=3, column=rating_col).value in (None, "")
        assert ws.cell(row=3, column=critique_col).value in (None, "")

    def test_filters_passed_to_get_contexts(self):
        """Filters dict is passed through to the get_contexts function."""
        received_filters = {}
        def mock_get_contexts(f):
            received_filters.update(f)
            return {"items": []}
        from grader.batch import export_to_excel
        export_to_excel(
            {"status": "reviewed", "flagged": "true"},
            _get_contexts=mock_get_contexts,
            _get_context_data=lambda cid: None,
            _get_review=lambda cid: None,
        )
        assert received_filters == {"status": "reviewed", "flagged": "true"}

    def test_option_columns_populated(self):
        """option_A through option_D are written from the question options dict."""
        ctx_data = {
            "context_id": "ctx-4",
            "passage": "Passage",
            "questions": [{
                "question_id": 1,
                "correct_answer": "B",
                "grammar_topic": "tense",
                "options": {"A": "alpha", "B": "beta", "C": "gamma", "D": "delta"},
                "explanation": {"why_correct": "W", "grammar_rule": "R"},
            }],
        }
        file_bytes = self._run_export(
            {"items": [{"context_id": "ctx-4"}]},
            {"ctx-4": ctx_data},
            {},
        )
        wb = _load_wb(file_bytes)
        ws = wb.active
        from grader.batch import COLUMNS
        assert ws.cell(row=2, column=COLUMNS.index("option_A") + 1).value == "alpha"
        assert ws.cell(row=2, column=COLUMNS.index("option_B") + 1).value == "beta"
        assert ws.cell(row=2, column=COLUMNS.index("option_C") + 1).value == "gamma"
        assert ws.cell(row=2, column=COLUMNS.index("option_D") + 1).value == "delta"


# ── Import unit tests ─────────────────────────────────────────────────────────

class TestImportFromExcel:
    def _run_import(self, rows, save_calls=None):
        from grader.batch import import_from_excel
        calls = save_calls if save_calls is not None else []

        def mock_save(context_id, expert_rating, expert_critique):
            calls.append((context_id, expert_rating, expert_critique))
            return {"updated_at": "2026-01-01T00:00:00"}

        file_bytes = _make_xlsx(rows)
        return import_from_excel(file_bytes, _save_review=mock_save), calls

    def test_valid_good_rating_saved(self):
        """A row with rating 'Good' calls save_review with correct args."""
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            ["ctx-abc", "Good", "Looks fine."],
        ]
        result, calls = self._run_import(rows)
        assert result["imported"] == 1
        assert result["skipped"] == 0
        assert result["errors"] == []
        assert calls == [("ctx-abc", "Good", "Looks fine.")]

    def test_valid_bad_rating_saved(self):
        """A row with rating 'Bad' calls save_review."""
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            ["ctx-xyz", "Bad", "Has errors"],
        ]
        result, calls = self._run_import(rows)
        assert result["imported"] == 1
        assert calls == [("ctx-xyz", "Bad", "Has errors")]

    def test_blank_rating_skipped(self):
        """Row with blank expert_rating is counted as skipped."""
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            ["ctx-1", "", ""],
            ["ctx-2", "Good", "OK"],
        ]
        result, calls = self._run_import(rows)
        assert result["imported"] == 1
        assert result["skipped"] == 1
        assert len(calls) == 1

    def test_invalid_rating_added_to_errors(self):
        """Row with rating 'Maybe' goes to errors; subsequent rows still processed."""
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            ["ctx-bad", "Maybe", ""],
            ["ctx-good", "Good", "Fine"],
        ]
        result, calls = self._run_import(rows)
        assert result["imported"] == 1
        assert result["errors"] == [{"context_id": "ctx-bad", "reason": "Invalid expert_rating: 'Maybe'"}]
        assert len(calls) == 1

    def test_unknown_context_id_added_to_errors(self):
        """save_review returns None for unknown context → added to errors."""
        from grader.batch import import_from_excel
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            ["unknown-ctx", "Good", ""],
        ]

        def mock_save_none(context_id, expert_rating, expert_critique):
            return None

        file_bytes = _make_xlsx(rows)
        result = import_from_excel(file_bytes, _save_review=mock_save_none)
        assert result["imported"] == 0
        assert result["errors"] == [{"context_id": "unknown-ctx", "reason": "Context not found in database"}]

    def test_contiguous_block_only_first_row_processed(self):
        """Multiple rows for same context_id — only first row's rating is used."""
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            ["ctx-1", "Good", "First row"],
            ["ctx-1", "Bad", "Second row — should be ignored"],
            ["ctx-1", "Bad", "Third row — should be ignored"],
        ]
        result, calls = self._run_import(rows)
        assert result["imported"] == 1
        assert calls == [("ctx-1", "Good", "First row")]

    def test_missing_required_column_raises(self):
        """File without expert_rating column raises ValueError with message."""
        from grader.batch import import_from_excel
        rows = [
            ["context_id", "passage"],
            ["ctx-1", "Some text"],
        ]
        file_bytes = _make_xlsx(rows)
        with pytest.raises(ValueError, match="Missing required columns"):
            import_from_excel(file_bytes)

    def test_non_xlsx_raises(self):
        """Non-xlsx bytes raise ValueError."""
        from grader.batch import import_from_excel
        with pytest.raises(ValueError, match="Invalid file format"):
            import_from_excel(b"this is not an xlsx file")

    def test_case_insensitive_column_matching(self):
        """Column names matched case-insensitively."""
        rows = [
            ["Context_ID", "Expert_Rating", "Expert_Critique"],
            ["ctx-1", "Good", "OK"],
        ]
        result, calls = self._run_import(rows)
        assert result["imported"] == 1
        assert calls[0][0] == "ctx-1"

    def test_whitespace_stripped_column_names(self):
        """Column names with leading/trailing spaces still match."""
        rows = [
            ["  context_id  ", "  expert_rating  ", "  expert_critique  "],
            ["ctx-1", "Good", "OK"],
        ]
        result, calls = self._run_import(rows)
        assert result["imported"] == 1

    def test_null_critique_treated_as_none(self):
        """Blank expert_critique is passed as None to save_review."""
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            ["ctx-1", "Good", ""],
        ]
        result, calls = self._run_import(rows)
        assert calls[0][2] is None


# ── Integration fixtures ──────────────────────────────────────────────────────

@pytest.fixture
def db_path(tmp_path):
    """Provide a temporary database path and patch DB_PATH in both modules."""
    path = str(tmp_path / "test_batch.db")
    import tools.question_bank as qb
    import tools.grader_db as gdb
    old_qb, old_gdb = qb.DB_PATH, gdb.DB_PATH
    qb.DB_PATH = path
    gdb.DB_PATH = path
    yield path
    qb.DB_PATH = old_qb
    gdb.DB_PATH = old_gdb


@pytest.fixture
def client(db_path):
    """Flask test client backed by a temporary database."""
    from grader.app import create_app
    app = create_app()
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


def _seed_contexts(db_path: str, count: int) -> list:
    """Insert test contexts and return their context_ids."""
    from tools.question_bank import init_db
    from tools.grader_db import init_reviews_table
    init_db()
    init_reviews_table()

    now = datetime.now().isoformat()
    questions = [{
        "question_id": 1,
        "options": {"A": "a", "B": "b", "C": "c", "D": "d"},
        "correct_answer": "A",
        "grammar_topic": "agreement",
        "explanation": {"why_correct": "Reason", "grammar_rule": "Rule"},
    }]
    questions_json = json.dumps(questions, ensure_ascii=False)

    context_ids = []
    conn = sqlite3.connect(db_path)
    try:
        for i in range(count):
            cid = str(uuid.uuid4())
            conn.execute(
                """INSERT INTO contexts
                   (context_id, type, passage, questions_json, num_questions,
                    grammar_topics, status, source_session, created_at,
                    times_served, passage_hash, last_incorrect, user_flags)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    cid, "fill_in_blank",
                    f"Test passage {i} with (1) _______________ blank.",
                    questions_json, 1, "agreement", "reviewed",
                    "test_session", now, 0, f"hash_{i}_{cid}", 0, 0,
                ),
            )
            context_ids.append(cid)
        conn.commit()
    finally:
        conn.close()
    return context_ids


# ── Export integration tests ──────────────────────────────────────────────────

class TestExportRoute:
    def test_export_returns_xlsx_content_type(self, client, db_path):
        """GET /api/export returns spreadsheetml content type."""
        _seed_contexts(db_path, 1)
        resp = client.get("/api/export")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type

    def test_export_returns_valid_xlsx(self, client, db_path):
        """GET /api/export returns a parseable .xlsx workbook with correct header."""
        _seed_contexts(db_path, 2)
        resp = client.get("/api/export")
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        from grader.batch import COLUMNS
        header = [ws.cell(row=1, column=i).value for i in range(1, len(COLUMNS) + 1)]
        assert header == COLUMNS

    def test_export_respects_filters(self, client, db_path):
        """GET /api/export?status=battle_tested returns only matching contexts."""
        context_ids = _seed_contexts(db_path, 3)
        conn = sqlite3.connect(db_path)
        conn.execute("UPDATE contexts SET status = 'battle_tested' WHERE context_id = ?", (context_ids[0],))
        conn.commit()
        conn.close()

        resp = client.get("/api/export?status=battle_tested")
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        from grader.batch import COLUMNS
        ctx_col = COLUMNS.index("context_id") + 1
        data_rows = [
            ws.cell(row=r, column=ctx_col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=ctx_col).value
        ]
        assert len(set(data_rows)) == 1
        assert data_rows[0] == context_ids[0]

    def test_export_empty_result_returns_header_only(self, client, db_path):
        """GET /api/export with no matching contexts returns header-only xlsx (not 404)."""
        _seed_contexts(db_path, 0)
        resp = client.get("/api/export")
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        assert ws.cell(row=2, column=1).value is None

    def test_export_content_disposition_header(self, client, db_path):
        """GET /api/export sets Content-Disposition: attachment with .xlsx filename."""
        _seed_contexts(db_path, 0)
        resp = client.get("/api/export")
        assert "attachment" in resp.headers.get("Content-Disposition", "")
        assert ".xlsx" in resp.headers.get("Content-Disposition", "")


# ── Import integration tests ──────────────────────────────────────────────────

class TestImportRoute:
    def _upload(self, client, file_bytes, filename="review.xlsx"):
        return client.post(
            "/api/import",
            data={"file": (io.BytesIO(file_bytes), filename)},
            content_type="multipart/form-data",
        )

    def test_import_valid_file_returns_counts(self, client, db_path):
        """POST /api/import with valid xlsx returns imported/skipped/errors."""
        context_ids = _seed_contexts(db_path, 1)
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            [context_ids[0], "Good", "Looks fine."],
        ]
        resp = self._upload(client, _make_xlsx(rows))
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["imported"] == 1
        assert data["skipped"] == 0
        assert data["errors"] == []

    def test_import_no_file_returns_400(self, client, db_path):
        """POST /api/import with no file attachment returns 400."""
        _seed_contexts(db_path, 0)
        resp = client.post("/api/import", data={}, content_type="multipart/form-data")
        assert resp.status_code == 400
        assert "error" in resp.get_json()

    def test_import_non_xlsx_returns_400(self, client, db_path):
        """POST /api/import with a .csv file returns 400."""
        _seed_contexts(db_path, 0)
        resp = client.post(
            "/api/import",
            data={"file": (io.BytesIO(b"context_id,expert_rating\nctx-1,Good"), "review.csv")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 400
        data = resp.get_json()
        assert data["error"] == "Invalid file format"

    def test_import_missing_columns_returns_400(self, client, db_path):
        """POST /api/import with xlsx missing expert_rating column returns 400."""
        _seed_contexts(db_path, 0)
        rows = [["context_id", "passage"], ["ctx-1", "text"]]
        resp = self._upload(client, _make_xlsx(rows))
        assert resp.status_code == 400
        data = resp.get_json()
        assert "Missing required columns" in data["error"]

    def test_import_updates_review_in_db(self, client, db_path):
        """Imported review is persisted — visible via GET /api/contexts/<id>."""
        context_ids = _seed_contexts(db_path, 1)
        cid = context_ids[0]
        rows = [
            ["context_id", "expert_rating", "expert_critique"],
            [cid, "Bad", "Found issues"],
        ]
        self._upload(client, _make_xlsx(rows))
        detail = client.get(f"/api/contexts/{cid}").get_json()
        assert detail["review"]["expert_rating"] == "Bad"
        assert detail["review"]["expert_critique"] == "Found issues"
