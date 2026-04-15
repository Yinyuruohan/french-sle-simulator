"""
LLM Grader — Batch Excel Export/Import

Provides export_to_excel() and import_from_excel() for bulk review operations.
COLUMNS is the single source of truth for column schema used by both functions.
"""
import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Column schema — order determines column position in the Excel file
COLUMNS = [
    "context_id",
    "passage",
    "correct_answer",
    "grammar_topic",
    "option_A",
    "option_B",
    "option_C",
    "option_D",
    "why_correct",
    "grammar_rule",
    "expert_rating",    # editable — first question row of each context only
    "expert_critique",  # editable — first question row of each context only
]

EDITABLE_COLS = {"expert_rating", "expert_critique"}

COL_WIDTHS = {
    "passage": 80,
    "expert_critique": 50,
    "context_id": 36,
    "why_correct": 50,
    "grammar_rule": 50,
    "option_A": 30,
    "option_B": 30,
    "option_C": 30,
    "option_D": 30,
}
DEFAULT_WIDTH = 20

_YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_GREY = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")


def export_to_excel(
    filters: dict,
    *,
    _get_contexts=None,
    _get_context_data=None,
    _get_review=None,
) -> bytes:
    """
    Build an Excel workbook from contexts matching filters.

    Returns bytes of the .xlsx file.
    One row per question; context-level fields (context_id, passage) repeat.
    expert_rating and expert_critique are editable (yellow, unlocked) on the
    first question row of each context only. All other cells are greyed and locked.
    Header row is frozen at row 2.
    """
    from tools.grader_db import get_contexts_for_review, get_context_data, get_review

    get_contexts = _get_contexts or get_contexts_for_review
    get_ctx_data = _get_context_data or get_context_data
    get_rev = _get_review or get_review

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expert Review"

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = _GREY
        cell.protection = Protection(locked=True)

    ws.freeze_panes = "A2"

    # ── Column widths ───────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, DEFAULT_WIDTH)

    # ── Data rows ───────────────────────────────────────────────────────────
    result = get_contexts(filters)
    row_idx = 2

    for ctx_item in result.get("items", []):
        context_id = ctx_item["context_id"]
        ctx_data = get_ctx_data(context_id)
        review = get_rev(context_id)

        if ctx_data is None:
            continue

        expert_rating = (review.get("expert_rating") if review else None) or ""
        expert_critique = (review.get("expert_critique") if review else None) or ""
        questions = ctx_data.get("questions", [])

        for q_idx, q in enumerate(questions):
            is_first = q_idx == 0
            expl = q.get("explanation") or {}
            opts = q.get("options") or {}

            row_data = {
                "context_id": context_id,
                "passage": ctx_data.get("passage", ""),
                "correct_answer": q.get("correct_answer", ""),
                "grammar_topic": q.get("grammar_topic", ""),
                "option_A": opts.get("A", ""),
                "option_B": opts.get("B", ""),
                "option_C": opts.get("C", ""),
                "option_D": opts.get("D", ""),
                "why_correct": expl.get("why_correct", ""),
                "grammar_rule": expl.get("grammar_rule", ""),
                "expert_rating": expert_rating if is_first else "",
                "expert_critique": expert_critique if is_first else "",
            }

            for col_idx, col_name in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_name in EDITABLE_COLS and is_first:
                    cell.fill = _YELLOW
                    cell.protection = Protection(locked=False)
                else:
                    cell.fill = _GREY
                    cell.protection = Protection(locked=True)

            row_idx += 1

    # ── Dropdown validation for expert_rating ──────────────────────────────
    if row_idx > 2:
        rating_col = get_column_letter(COLUMNS.index("expert_rating") + 1)
        dv = DataValidation(
            type="list",
            formula1='"Good,Bad"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"{rating_col}2:{rating_col}{row_idx - 1}"
        ws.add_data_validation(dv)

    # Enable sheet protection — activates all cell-level locked=True settings
    ws.protection.sheet = True
    ws.protection.password = ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def import_from_excel(file_bytes: bytes, *, _save_review=None) -> dict:
    """
    Parse an uploaded .xlsx workbook and save expert reviews.

    Column names matched case-insensitively with leading/trailing whitespace stripped.
    Assumes contiguous context blocks: rows for the same context_id are consecutive.
    Only the first row of each context_id block provides expert_rating/expert_critique;
    subsequent rows in the same block are skipped.

    Returns {"imported": N, "skipped": M, "errors": [{"context_id": ..., "reason": ...}]}
    Raises ValueError for structural issues (invalid format, missing required columns).
    """
    from tools.grader_db import save_review
    save_rev = _save_review or save_review

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValueError("Invalid file format")

    ws = wb.active
    rows = list(ws.rows)

    if not rows:
        raise ValueError("Missing required columns: context_id, expert_rating")

    # Parse header: case-insensitive, whitespace-stripped
    header = [str(cell.value or "").strip().lower() for cell in rows[0]]
    col_index = {name: idx for idx, name in enumerate(header)}

    required = {"context_id", "expert_rating"}
    missing = required - set(col_index.keys())
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(sorted(missing)))

    def get_val(row, col_name):
        idx = col_index.get(col_name.lower())
        if idx is None:
            return ""
        val = row[idx].value
        return str(val).strip() if val is not None else ""

    imported = 0
    skipped = 0
    errors = []
    current_context_id = None

    for row in rows[1:]:
        context_id = get_val(row, "context_id")
        if not context_id:
            continue
        # Non-first row of a contiguous context block — skip review fields
        if context_id == current_context_id:
            continue

        current_context_id = context_id
        expert_rating = get_val(row, "expert_rating")
        expert_critique = get_val(row, "expert_critique") or None

        if not expert_rating:
            skipped += 1
            continue

        if expert_rating not in ("Good", "Bad"):
            errors.append({
                "context_id": context_id,
                "reason": f"Invalid expert_rating: {expert_rating!r}",
            })
            continue

        result = save_rev(context_id, expert_rating, expert_critique)
        if result is None:
            errors.append({"context_id": context_id, "reason": "Context not found in database"})
        else:
            imported += 1

    return {"imported": imported, "skipped": skipped, "errors": errors}